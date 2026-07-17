"""Regenerate src/domain/formOptions.ts from the workbook's "Details" sheet.

Usage (from the app/ directory):
    python scripts/generate-form-options.py

The Details sheet holds the form's flat option lists, one per column. The
columns are jagged (10 sites vs 18 parent assets vs 27 components), so row
position carries NO linkage between columns — Site→Parent Asset linkage comes
from the Asset Registry / assets cache, not from this sheet. The one row-wise
pairing that IS meaningful is Priority ↔ Condition (P1..P5 aligned to their
condition descriptions).
"""
import warnings, json, os, sys

warnings.filterwarnings("ignore")
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
WORKBOOK = os.path.join(HERE, "..", "..", "Nerizon_Structural_Integrity_Assessment_NALA.xlsx")
OUT = os.path.join(HERE, "..", "src", "domain", "formOptions.ts")

wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
ws = wb["Details"]

COLS = {
    "sites": 1,  # A: Site
    "parentAssets": 2,  # B: Parent Asset
    "categories": 3,  # C: Category
    "components": 4,  # D: Component
    "subComponents": 5,  # E: Sub-Component
    "locationDescriptions": 6,  # F: Location Description (provinces)
    "deficiencyCategories": 7,  # G: Deficiency Categories (15 labels)
    "deficiencyDescriptions": 8,  # H: Deficiency Description
}


def colvals(col):
    out = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is not None and str(v).strip() != "":
            out.append(str(v).strip())
    return out


lists = {name: colvals(col) for name, col in COLS.items()}

# Priority (I) ↔ Condition (J): row-aligned pairing
priorities = []
for r in range(2, ws.max_row + 1):
    p = ws.cell(row=r, column=9).value
    c = ws.cell(row=r, column=10).value
    if p and c:
        priorities.append({"priority": str(p).strip(), "condition": str(c).strip()})

for name, vals in lists.items():
    if len(set(vals)) != len(vals):
        sys.exit(f"ABORT: duplicate values in {name}: {sorted(set(v for v in vals if vals.count(v) > 1))}")

ts = []
ts.append('// GENERATED from Nerizon_Structural_Integrity_Assessment_NALA.xlsx ("Details" sheet).')
ts.append("// Regenerate with scripts/generate-form-options.py rather than editing by hand.")
ts.append("//")
ts.append("// These are FLAT option lists — the sheet's columns are jagged, so row position")
ts.append("// carries no linkage between columns. Site→Parent Asset linkage comes from the")
ts.append("// Asset Registry / assets cache. The one meaningful row pairing is Priority↔Condition.")
ts.append("")
ts.append("export interface PriorityCondition {")
ts.append("  priority: string;")
ts.append("  condition: string;")
ts.append("}")
ts.append("")
ts.append("export const FORM_OPTIONS = {")
for name in COLS:
    ts.append(f"  {name}: {json.dumps(lists[name], ensure_ascii=False)},")
ts.append(f"  priorityConditions: {json.dumps(priorities, ensure_ascii=False)},")
ts.append("} as const;")
ts.append("")
ts.append("export function conditionFor(priority: string): string {")
ts.append("  return FORM_OPTIONS.priorityConditions.find((p) => p.priority === priority)?.condition ?? '';")
ts.append("}")
ts.append("")

with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(ts))
print(f"written {OUT}: " + ", ".join(f"{k}={len(v)}" for k, v in lists.items()) + f", priorityConditions={len(priorities)}")

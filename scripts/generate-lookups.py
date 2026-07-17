"""Regenerate src/domain/lookups.ts from the source workbook.

Usage (from the app/ directory):
    python scripts/generate-lookups.py

Reads ../Nerizon_Structural_Integrity_Assessment_NALA.xlsx (hidden Lookups sheet,
Details, Deficiency Reference) and rewrites src/domain/lookups.ts.
"""
import warnings, json, re, os

warnings.filterwarnings("ignore")
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
WORKBOOK = os.path.join(HERE, "..", "..", "Nerizon_Structural_Integrity_Assessment_NALA.xlsx")
OUT = os.path.join(HERE, "..", "src", "domain", "lookups.ts")

wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
lk = wb["Lookups"]
det = wb["Details"]
ref = wb["Deficiency Reference"]


def colvals(ws, col, start=2):
    out = []
    for r in range(start, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is not None and str(v).strip() != "":
            out.append(str(v).strip())
    return out


hdr = {}
for j in range(1, lk.max_column + 1):
    h = lk.cell(row=1, column=j).value
    if h:
        hdr[str(h).strip()] = j

sites = colvals(det, 1)
categories = colvals(lk, hdr["DeficiencyCategory_List"])
component_types = colvals(lk, hdr["ComponentType_List"])
priorities = colvals(lk, hdr["Priority_List"])
risk = colvals(lk, hdr["Risk_List"])

prio_desc = {}
for r in range(2, det.max_row + 1):
    p = det.cell(row=r, column=9).value
    c = det.cell(row=r, column=10).value
    if p and c:
        prio_desc[str(p).strip()] = str(c).strip()

defs = {}
for r in range(2, ref.max_row + 1):
    code = ref.cell(row=r, column=1).value
    d = ref.cell(row=r, column=3).value
    if code:
        defs[str(code).strip()] = str(d).strip() if d else ""

cats = []
for label in categories:
    code = label.split(" ")[0].strip()

    def get(prefix):
        key = f"{prefix}_{code}"
        return colvals(lk, hdr[key]) if key in hdr else []

    cats.append(
        {
            "code": code,
            "label": label,
            "definition": defs.get(code, ""),
            "descriptions": get("Def"),
            "mechanisms": get("Mech"),
            "focusAreas": get("Focus"),
        }
    )

comp = []
for ct in component_types:
    k = "Sub_" + re.sub(r"[ &/]", "", ct)
    subs = colvals(lk, hdr[k]) if k in hdr else []
    comp.append({"name": ct, "subComponents": subs})

ts = []
ts.append("// GENERATED from Nerizon_Structural_Integrity_Assessment_NALA.xlsx (hidden Lookups sheet + Details + Deficiency Reference).")
ts.append("// Regenerate with scripts/generate-lookups.py rather than editing by hand.")
ts.append("")
ts.append("export interface DeficiencyCategory {")
ts.append("  code: string;")
ts.append("  label: string;")
ts.append("  definition: string;")
ts.append("  descriptions: string[];")
ts.append("  mechanisms: string[];")
ts.append("  focusAreas: string[];")
ts.append("}")
ts.append("")
ts.append("export interface ComponentType {")
ts.append("  name: string;")
ts.append("  subComponents: string[];")
ts.append("}")
ts.append("")
ts.append(f"export const SITES: string[] = {json.dumps(sites, ensure_ascii=False)};")
ts.append("")
ts.append(f"export const PRIORITIES = {json.dumps(priorities)} as const;")
ts.append("export type PriorityRating = (typeof PRIORITIES)[number];")
ts.append("")
ts.append(f"export const PRIORITY_DESCRIPTIONS: Record<string, string> = {json.dumps(prio_desc, ensure_ascii=False, indent=2)};")
ts.append("")
ts.append(f"export const RISK_LEVELS = {json.dumps(risk)} as const;")
ts.append("export type RiskLevel = (typeof RISK_LEVELS)[number];")
ts.append("")
ts.append(f"export const DEFICIENCY_CATEGORIES: DeficiencyCategory[] = {json.dumps(cats, ensure_ascii=False, indent=2)};")
ts.append("")
ts.append(f"export const COMPONENT_TYPES: ComponentType[] = {json.dumps(comp, ensure_ascii=False, indent=2)};")
ts.append("")

with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(ts))
print(f"written {OUT}: {len(sites)} sites, {len(cats)} categories, {len(comp)} component types")
